from time import sleep

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver
from difflib import SequenceMatcher
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

path = 'Permits-Issued-to-Companies-2021.xlsx'
workbook = load_workbook(path)
sheet = workbook['Sheet1']


def main():
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    driver.get('https://www.linkedin.com/uas/login')
    driver.maximize_window()
    login(driver)


def login(driver):
    username = driver.find_element_by_id('username')
    username_text = input("Please enter username: ")
    username.send_keys(username_text)
    password = driver.find_element_by_id('password')
    password_text = input("Please enter password: ")
    password.send_keys(password_text)
    log_in_button = driver.find_element_by_class_name('from__button--floating')
    log_in_button.click()
    start(driver)


def start(driver):
    iniciar = True
    row = 5
    while iniciar:
        value = sheet.cell(row=row, column=1).value
        if value is not None:
            search(driver, value.strip(), row)
            row = row + 1
        else:
            iniciar = False


def search(driver, name, row):
    try:
        input_search = driver.find_element_by_class_name('search-global-typeahead__input')
    except Exception as e:
        # Tela de validação do Linkedin, seja rápido rs...
        sleep(25)
        input_search = driver.find_element_by_class_name('search-global-typeahead__input')
    input_search.clear()
    input_search.send_keys(name)
    input_search.send_keys(Keys.ENTER)
    sleep(3)
    url = driver.current_url
    url = url.replace("/search/results/all", "/search/results/companies")
    try:
        driver.get(url)
    except requests.Timeout as err:
        print("###### ERRO ######")
        print(err)
        print("##################")
        driver.refresh()
        search(driver, name, row + 1)
    select_company(driver, name, row)


def select_company(driver, name, row):
    html_page = driver.page_source
    soup = BeautifulSoup(html_page, 'html.parser')

    empresas = soup.findAll('span', {'class': ['entity-result__title-text']})
    setor = soup.findAll('div', {'class': ['entity-result__primary-subtitle']})

    nome_link_dict = dict()
    for item in zip(empresas, setor):
        nome_link_dict[item[0].text.strip().title()] = [item[0].contents[1].attrs['href'], item[1].text.strip().split(" • ")[0]]

    nome_empresa = None
    is_true = False

    if nome_link_dict:
        for key, value in nome_link_dict.items():
            if name.upper() == key.upper():
                nome_empresa = key.title()
                is_true = True
                break
        if not is_true:
            for key, value in nome_link_dict.items():
                porcentagem = similaridade(name.upper(), key.upper())
                if porcentagem >= 80:
                    nome_empresa = key.title()
                    print(f'[{row}] % {porcentagem} de similaridade.')
                    break
        if nome_empresa is not None:
            url = nome_link_dict[nome_empresa][0]
            setor = nome_link_dict[nome_empresa][1]
            get_information(row, url, setor, name)
        else:
            print(f'[{row}] {name} - Não encontrado na busca')
    else:
        print(f'[{row}] {name} - Não existe no Linkedin')


def similaridade(name_xlsx, name_scraping):
    seq = SequenceMatcher(None, name_xlsx, name_scraping)
    porcentagem = seq.ratio() * 100
    return round(porcentagem, 2)


def get_information(row, url, setor, name):
    sheet.cell(row=row, column=11).value = setor
    sheet.cell(row=row, column=12).value = url
    workbook.save(path)
    print(f'[{row}] {name} - {setor} - {url}!')


if __name__ == '__main__':
    main()
