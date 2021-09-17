from time import sleep
from difflib import SequenceMatcher
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

path = 'Permits-Issued-to-Companies-2021.xlsx'
workbook = load_workbook(path)
sheet = workbook['Sheet1']


def main():
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    driver.get('https://bi.experian.ie/scripts/cgiip.exe/WService=EXPWeb/Online/OnlineSearch.w')
    driver.maximize_window()
    driver = select_new_aba(driver)
    driver.switch_to.frame(driver.find_elements_by_tag_name("frame")[0])
    sleep(5)
    start(driver)


def select_new_aba(driver):
    input_search = driver.find_element_by_id('Name')
    input_search.clear()
    input_search.send_keys(Keys.ENTER)
    driver.switch_to.window(driver.window_handles[0])
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    return driver


def start(driver):
    iniciar = True
    row = 5
    while iniciar:
        value = sheet.cell(row=row, column=1).value
        if value is not None:
            search(driver, value.strip(), row)
            row = row + 1
        else:
            iniciar = False


def search(driver, name, row):
    html_page = driver.page_source
    soup = BeautifulSoup(html_page, 'html.parser')
    soup.findAll('div', {'class': ['resultDetails']})
    try:
        input_search = driver.find_element_by_id('Name')
        input_search.clear()
        input_search.send_keys(name)
        input_search.send_keys(Keys.ENTER)
        scraping(driver, name, row)
    except Exception as e:
        print(f'Erro com a empresa: {name}')
        driver.refresh()
        sleep(3)
        driver.switch_to.frame(driver.find_elements_by_tag_name("frame")[0])
        search(driver, name, row)


def scraping(driver, name, row):
    html_page = driver.page_source
    soup = BeautifulSoup(html_page, 'html.parser')
    empresas = soup.findAll('div', {'class': ['resultDetails']})
    nome_link_dict = {item.contents[1].text.strip(): item.contents[4].text.split("Category:")[1].strip() for item in
                      empresas}
    is_save = False
    if nome_link_dict:
        for key, value in nome_link_dict.items():
            if name.upper() == key.upper():
                is_save = save(row, value, name)
                break
        if not is_save:
            for key, value in nome_link_dict.items():
                porcentagem = similaridade(name.upper(), key.upper())
                if porcentagem > 80:
                    save(row, value, name)
                    print(f'%{porcentagem} de similaridade.')
                    break
    else:
        print(f'[{row}] {name} - Não encontrado na busca')


def save(row, value, name):
    sheet.cell(row=row, column=13).value = value
    workbook.save(path)
    print(f'[{row}] {name} - {value}')
    return True


def similaridade(name_xlsx, name_scraping):
    seq = SequenceMatcher(None, name_xlsx, name_scraping)
    porcentagem = seq.ratio() * 100
    return round(porcentagem, 2)


if __name__ == '__main__':
    main()
